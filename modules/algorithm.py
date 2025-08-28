import cv2
import numpy as np
import os
import datetime



class Algorithm:
    def __init__(self, main_window=None):

        self.image_path = None
        self.data = []
        self.RNA_type = None
        self.coefficients = None

        if main_window is not None:
            self.main_window = main_window
            self.rows = self.main_window.rows
            self.cols = self.main_window.cols
            self.r = self.main_window.r
            self.precision = self.main_window.precision

            self.argu_update_rna_type()

        else:
            self.rows = 4
            self.cols = 6
            self.r = 0.2
            self.precision = 4
    def save_detection_data(self, output_dir, original_image_path):
        """
        保存检测参数和原始图片到输出文件夹，并生成记录参数的Excel表格
        """
        # 创建输出目录
        os.makedirs(output_dir, exist_ok=True)
        cv2.imwrite(os.path.join(output_dir, 'origin_img.jpg'), self.main_window.origin_img)
        
        # 生成参数信息
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = "检测参数"
            
            # 创建居中对齐样式
            center_alignment = Alignment(horizontal="center", vertical="center")
            bold_font = Font(bold=True)
            
            # 写入参数信息
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 100
            
            # 标题行
            ws.cell(row=1, column=1, value="参数名称").font = bold_font
            ws.cell(row=1, column=1).alignment = center_alignment
            ws.cell(row=1, column=2, value="参数值").font = bold_font
            ws.cell(row=1, column=2).alignment = center_alignment
            
            # 写入具体参数
            row_index = 2
            
            # 行列数参数
            ws.cell(row=row_index, column=1, value="行数").alignment = center_alignment
            ws.cell(row=row_index, column=2, value=self.rows).alignment = center_alignment
            row_index += 1
            
            ws.cell(row=row_index, column=1, value="列数").alignment = center_alignment
            ws.cell(row=row_index, column=2, value=self.cols).alignment = center_alignment
            row_index += 1
            
            ws.cell(row=row_index, column=1, value="半径").alignment = center_alignment
            ws.cell(row=row_index, column=2, value=self.r).alignment = center_alignment
            row_index += 1
            
            ws.cell(row=row_index, column=1, value="精度").alignment = center_alignment
            ws.cell(row=row_index, column=2, value=self.precision).alignment = center_alignment
            row_index += 1
            
            # RNA类型和拟合曲线
            ws.cell(row=row_index, column=1, value="类型").alignment = center_alignment
            ws.cell(row=row_index, column=2, value=self.RNA_type if self.RNA_type else "未指定").alignment = center_alignment
            row_index += 1
            
            # 生成拟合曲线表达式
            if hasattr(self, 'coefficients') and self.coefficients is not None:
                # 格式化系数，保留三位有效数字
                coeffs_formatted = [f"{coef:.3g}" for coef in self.coefficients]
                terms = []
                for i, coef in enumerate(coeffs_formatted):
                    power = len(coeffs_formatted) - 1 - i
                    if power == 0:
                        terms.append(f"{coef}")
                    elif power == 1:
                        terms.append(f"{coef}*x")
                    else:
                        terms.append(f"{coef}*x^{power}")
                
                expression = "P(x) = " + " + ".join(terms).replace("+ -", "- ")
            else:
                expression = "未生成拟合曲线"
                
            ws.cell(row=row_index, column=1, value="曲线").alignment = center_alignment
            ws.cell(row=row_index, column=2, value=expression).alignment = center_alignment
            
            # 保存Excel文件
            wb.save(os.path.join(output_dir, 'parameters.xlsx'))
            print(f"检测参数已保存至: {os.path.join(output_dir, 'parameters.xlsx')}")
            
        except Exception as e:
            print(f"保存检测参数失败: {str(e)}")
    def save_results(self, output_dir, stitched_img, sub_image_info):
        """
        保存处理结果，包括图片和Excel数据
        """
        os.makedirs(output_dir, exist_ok=True)
        cv2.imwrite(os.path.join(output_dir, 'result_img.jpg'), stitched_img)
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
            from openpyxl.drawing.image import Image as XLImage
            from PIL import Image
            import io
            import math
            
            wb = Workbook()
            
            # 创建一个包含所有数据的工作表
            ws1 = wb.active
            ws1.title = "详细检测数据"
            
            # 创建居中对齐样式
            center_alignment = Alignment(horizontal="center", vertical="center")
            
            # 写入表头
            headers = ['行号', '列号', '原始灰度', '相对灰度', '预测浓度', 'lg(浓度)', '原始子图']
            for col, header in enumerate(headers, 1):
                cell = ws1.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = center_alignment
            
            # 写入详细数据 - 按照您的要求，按行优先顺序排列
            for idx, info in enumerate(sub_image_info):
                # 每一行的数据连续排列，然后换行
                row = idx + 2  # 第一行是表头，所以数据从第二行开始
                
                # 计算log10(浓度)
                density = info['density']
                log_density = math.log10(density) if density > 0 else 0
                
                # 写入数据并居中对齐
                ws1.cell(row=row, column=1, value=info['i'] + 1).alignment = center_alignment  # 行号（从1开始）
                ws1.cell(row=row, column=2, value=info['j'] + 1).alignment = center_alignment  # 列号（从1开始）
                ws1.cell(row=row, column=3, value=info['avg_gray']).alignment = center_alignment  # 原始灰度值
                ws1.cell(row=row, column=4, value=info['gray_diff']).alignment = center_alignment  # 相对灰度值
                ws1.cell(row=row, column=5, value=info['density']).alignment = center_alignment  # 预测浓度
                ws1.cell(row=row, column=6, value=log_density).alignment = center_alignment  # log10(浓度)

                # 插入子图
                sub_image = info['sub_image']
                if sub_image is not None:
                    # 将OpenCV图像转换为PIL图像
                    sub_image_rgb = cv2.cvtColor(sub_image, cv2.COLOR_BGR2RGB)
                    pil_image = Image.fromarray(sub_image_rgb)
                    
                    # 将PIL图像保存到内存中的字节流
                    image_stream = io.BytesIO()
                    pil_image.save(image_stream, format='PNG')
                    image_stream.seek(0)
                    
                    # 创建Excel图像对象
                    img = XLImage(image_stream)
                    
                    # 调整图像大小
                    img.width = 80
                    img.height = 80
                    
                    # 插入图像到单元格
                    # 使用标准方法添加图片到指定单元格
                    img.anchor = f'G{row}'  # G列是原始子图列
                    ws1.add_image(img)

            # 设置列宽
            ws1.column_dimensions['A'].width = 8   # "行号"列宽为5字符
            ws1.column_dimensions['B'].width = 8   # "列号"列宽为5字符
            ws1.column_dimensions['C'].width = 12  # "原始灰度"列宽为12字符
            ws1.column_dimensions['D'].width = 12  # "相对灰度"列宽为12字符
            ws1.column_dimensions['E'].width = 16  # "预测浓度"列宽为15字符
            ws1.column_dimensions['F'].width = 16  # "log10(浓度)"列宽为12字符
            ws1.column_dimensions['G'].width = 12  # "原始子图"列宽为12字符
            
            # 设置行高
            ws1.row_dimensions[1].height = 30  # 第一行表头行高为30
            for row in range(2, len(sub_image_info) + 2):  # 数据行
                ws1.row_dimensions[row].height = 60  # 设置行高以便显示图像
            
            wb.save(os.path.join(output_dir, 'results.xlsx'))
            print(f"Excel已保存至: {os.path.join(output_dir, 'results.xlsx')}")
        except Exception as e:
            print(f"Excel保存失败: {str(e)}")
        
        # 关闭预览窗口
        if hasattr(self, 'preview_window'):
            self.preview_window.close()
            
        # 询问是否打开输出目录
        try:
            from PySide6.QtWidgets import QMessageBox
            msg_box = QMessageBox()
            msg_box.setWindowTitle("保存完成")
            msg_box.setText("结果已保存成功！")
            msg_box.setInformativeText("是否要打开输出目录？")
            msg_box.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            msg_box.setDefaultButton(QMessageBox.Yes)
            
            result = msg_box.exec()
            if result == QMessageBox.Yes:
                # 打开目录
                import subprocess
                import platform
                system = platform.system()
                if system == "Windows":
                    subprocess.Popen(f'explorer "{os.path.abspath(output_dir)}"')
                elif system == "Darwin":  # macOS
                    subprocess.Popen(["open", os.path.abspath(output_dir)])
                else:  # Linux
                    subprocess.Popen(["xdg-open", os.path.abspath(output_dir)])
        except Exception as e:
            print(f"打开目录时出错: {e}")


    def argu_update_rna_type(self):

        try:
            if hasattr(self.main_window.ui, 'type_Box'):
                self.RNA_type = self.main_window.ui.type_Box.currentText()
            else:
                raise AttributeError("UI中不存在'type_Box'控件")
        except Exception as e:
            print(f"错误: {e}")
            
        if self.RNA_type == 'miR-223':
            x = np.array([3.3, 12.33, 20.20, 45, 57.73, 64.67])
            y = np.array([0, 1, 2, 3, 4, 5])
        elif self.RNA_type == 'miR-935':
            x = np.array([1.8, 12.3, 31.5, 50.7, 60, 66.1])
            y = np.array([0, 1, 2, 3, 4, 5])
        elif self.RNA_type == 'miR-2284W':
            x = np.array([1.8, 17.27, 18.24, 52, 50.4, 60.5])
            y = np.array([0, 1, 2, 3, 4, 5])

        print(self.RNA_type)
        
        # 拟合标准曲线（最后的参数为需要拟合的次数）
        self.coefficients = np.polyfit(x, y, 5)

        # 格式化系数，保留三位有效数字
        coeffs_formatted = [f"{coef:.3g}" for coef in self.coefficients]
        print("P(x) = ", end="")
        terms = []
        for i, coef in enumerate(coeffs_formatted):
            power = 5 - i
            if power == 0:
                terms.append(f"{coef}")
            elif power == 1:
                terms.append(f"{coef}*x")
            else:
                terms.append(f"{coef}*x^{power}")
            
        print(" + ".join(terms).replace("+ -", "- "))
    def count(self, main_window=None):
        # 初始化参数
        if main_window is not None:
            self.rows = self.main_window.rows
            self.cols = self.main_window.cols
            self.r = self.main_window.r
            self.precision = self.main_window.precision
            image = self.main_window.origin_img.copy()
        else:
            image = cv2.imread(self.image_path)

        height, width = image.shape[:2]
        row_height = height // self.rows
        col_width = width // self.cols

        # 绘制基础网格线
        for i in range(1, self.rows):
            cv2.line(image, (0, i*row_height), (width, i*row_height), (255,0,0), 2)
        for j in range(1, self.cols):
            cv2.line(image, (j*col_width, 0), (j*col_width, height), (255,0,0), 2)

        # 初始化拼接图像参数
        sub_h = int(row_height * 2 * self.r)
        sub_w = int(col_width * 2 * self.r)
        stitched_img = np.full((sub_h*self.rows, sub_w*self.cols, 3), 255, dtype=np.uint8)
        self.data = []

        # 存储每个子图的信息
        sub_image_info = []

        # ========== 第一次循环：计算所有子图的灰度值并存储 ==========
        for i in range(self.rows):
            for j in range(self.cols):
                # 计算中心坐标和采样区域
                cx = j * col_width + col_width // 2
                cy = i * row_height + row_height // 2
                radius = int(min(row_height, col_width) * self.r)

                # 裁剪子图
                start_row = max(cy - radius, 0)
                end_row = min(cy + radius, height)
                start_col = max(cx - radius, 0)
                end_col = min(cx + radius, width)
                sub_image = image[start_row:end_row, start_col:end_col]

                # 调整子图尺寸
                sub_image = cv2.resize(sub_image, (sub_w, sub_h))

                # 计算灰度值
                gray = cv2.cvtColor(sub_image, cv2.COLOR_BGR2GRAY)
                avg_gray = float(f"{cv2.mean(gray)[0]:.{self.precision}g}")

                # 保存子图信息
                sub_image_info.append({
                    'i': i,
                    'j': j,
                    'cx': cx,
                    'cy': cy,
                    'sub_image': sub_image.copy(),
                    'avg_gray': avg_gray,
                    'box': (start_row, end_row, start_col, end_col)  # 原图采样框坐标
                })

        # ========== 找到最大灰度值 ==========
        max_gray = max(info['avg_gray'] for info in sub_image_info)

        # ========== 第二次循环：计算灰度变化量并标注 ==========
        for info in sub_image_info:
            i = info['i']
            j = info['j']
            cx = info['cx']
            cy = info['cy']
            sub_image = info['sub_image'].copy()
            avg_gray = info['avg_gray']

            # y（灰度），x（浓度）之间的表达式
            # y = 5.803*In(x) + 0.473

            # 计算灰度变化量
            gray_diff = max_gray - avg_gray

            # y（灰度），x（浓度）之间的表达式
            # density = np.exp((gray_diff - 0.473) / 5.803)

            if gray_diff !=  0:
                density = 10**np.polyval(self.coefficients, gray_diff)
            else:
                density = 0

            # print(f"相对灰度：{gray_diff}，预测浓度：{density:.2f}")

            # 构建多行文本
            import math

            # 将density转换为 a*10^b 的形式
            if density > 0:
                exponent = math.floor(math.log10(density))
                mantissa = density / (10 ** exponent)
                if exponent == 0:
                    density_str = f"{mantissa:.{self.precision-1}g}"
                else:
                    density_str = f"{mantissa:.{self.precision-1}g}*10^{exponent}"
            else:
                density_str = "0"

            text_lines = [
                f"Avg: {avg_gray:.{self.precision}g}",
                f"Diff: {gray_diff:.{self.precision}g}",
                f"Density: {density_str}"
            ]
            
            # 更新子图信息，添加相对灰度值和预测浓度
            info.update({
                'gray_diff': gray_diff,
                'density': density,
                'density_str': density_str
            })

            # ========== 子图文本绘制 ==========
            font = cv2.FONT_HERSHEY_SIMPLEX
            thickness = 1

            # 动态字体计算
            base_scale = 0.004
            font_scale = min(sub_w, sub_h) * base_scale
            font_scale = max(0.3, min(font_scale, 1.0))

            # 获取每行文本的尺寸
            line_heights = []
            line_widths = []
            for line in text_lines:
                (tw, th), baseline = cv2.getTextSize(line, font, font_scale, thickness)
                line_widths.append(tw)
                line_heights.append(th)

            max_line_width = max(line_widths)
            max_line_height = max(line_heights)

            # 自动换行处理
            max_width = sub_w * 0.8
            if max_line_width > max_width:
                font_scale *= max_width / max_line_width
                font_scale = max(0.2, min(font_scale, 1.0))

                # 重新计算尺寸
                line_heights = []
                line_widths = []
                for line in text_lines:
                    (tw, th), baseline = cv2.getTextSize(line, font, font_scale, thickness)
                    line_widths.append(tw)
                    line_heights.append(th)
                max_line_width = max(line_widths)
                max_line_height = max(line_heights)

            # 居中定位
            total_height = max_line_height * len(text_lines) + baseline * (len(text_lines) - 1)
            ty_base = (sub_h + total_height) // 2

            # 逐行绘制
            for idx, line in enumerate(text_lines):
                tx = (sub_w - cv2.getTextSize(line, font, font_scale, thickness)[0][0]) // 2
                ty = ty_base + idx * (max_line_height + baseline)
                cv2.putText(sub_image, line, (tx, ty),
                            font, font_scale, (0, 0, 0), thickness, cv2.LINE_AA)

            # 放入拼接图
            stitched_img[i * sub_h:(i + 1) * sub_h, j * sub_w:(j + 1) * sub_w] = sub_image

            # ========== 原始图像标注 ==========
            start_row, end_row, start_col, end_col = info['box']
            box_tl_x = int(cx - sub_w / 2)
            box_tl_y = int(cy - sub_h / 2)
            box_br_x = int(cx + sub_w / 2)
            box_br_y = int(cy + sub_h / 2)

            # 绘制采样框（蓝色）
            cv2.rectangle(image, (box_tl_x, box_tl_y),
                          (box_br_x, box_br_y), (0, 0, 255), 2)

            # 使用与子图一致的文本适配逻辑
            box_width = box_br_x - box_tl_x
            box_height = box_br_y - box_tl_y

            # 动态字体计算
            base_scale = 0.004
            text_scale = min(box_width, box_height) * base_scale
            text_scale = max(0.3, min(text_scale, 1.0))

            # 获取每行文本的尺寸
            line_heights = []
            line_widths = []
            for line in text_lines:
                (text_w, text_h), baseline = cv2.getTextSize(line, font, text_scale, thickness)
                line_widths.append(text_w)
                line_heights.append(text_h)

            max_line_width = max(line_widths)
            max_line_height = max(line_heights)

            # 自动换行处理
            max_text_width = box_width * 0.8
            if max_line_width > max_text_width:
                text_scale *= max_text_width / max_line_width
                text_scale = max(0.2, min(text_scale, 1.0))

                # 重新计算尺寸
                line_heights = []
                line_widths = []
                for line in text_lines:
                    (text_w, text_h), baseline = cv2.getTextSize(line, font, text_scale, thickness)
                    line_widths.append(text_w)
                    line_heights.append(text_h)
                max_line_width = max(line_widths)
                max_line_height = max(line_heights)

            # 高度二次验证
            max_text_height = box_height * 0.3
            if (max_line_height + baseline) * len(text_lines) > max_text_height:
                text_scale *= max_text_height / ((max_line_height + baseline) * len(text_lines))
                text_scale = max(0.2, min(text_scale, 1.0))

            # 精确居中定位
            total_text_height = (max_line_height + baseline) * len(text_lines) - baseline
            text_y_base = box_tl_y + (box_height + total_text_height) // 2

            # 边界保护
            text_y_base = min(box_br_y - baseline - 5, max(box_tl_y + max_line_height + 5, text_y_base))

            # 逐行绘制
            for idx, line in enumerate(text_lines):
                text_x = box_tl_x + (box_width - cv2.getTextSize(line, font, text_scale, thickness)[0][0]) // 2
                text_y = text_y_base + idx * (max_line_height + baseline)
                cv2.putText(image, line, (text_x, text_y),
                            font, text_scale, (0, 0, 255), int(thickness * 1.5), cv2.LINE_AA)


            # ========== 结果保存和显示 ==========
            timestamp = datetime.datetime.now().strftime('%Y%m%d-%H%M%S')
            output_dir = os.path.join('outputs', timestamp)
            
            # 弹窗询问是否保存结果
            try:
                from PySide6.QtWidgets import QMessageBox, QLabel, QVBoxLayout, QWidget, QPushButton, QHBoxLayout
                from PySide6.QtGui import QPixmap, QImage
                from PySide6.QtCore import Qt
                
                # 创建预览窗口
                preview_window = QWidget()
                preview_window.setWindowTitle('结果预览')
                preview_window.resize(800, 650)
                
                # 创建布局
                layout = QVBoxLayout()
                
                # 创建标签用于显示图片
                label = QLabel()
                label.setAlignment(Qt.AlignCenter)
                
                # 转换拼接图像以进行预览
                if stitched_img is not None:
                    # 转换颜色格式 (OpenCV是BGR,需要转换为RGB)
                    img_rgb = cv2.cvtColor(stitched_img, cv2.COLOR_BGR2RGB)
                    
                    # 转换为QImage
                    h, w, ch = img_rgb.shape
                    bytes_per_line = ch * w
                    q_img = QImage(img_rgb.data, w, h, bytes_per_line, QImage.Format_RGB888)
                    
                    # 转换为QPixmap并适应窗口大小
                    pixmap = QPixmap.fromImage(q_img)
                    label.setPixmap(pixmap.scaled(780, 500, Qt.KeepAspectRatio, Qt.SmoothTransformation))
                
                layout.addWidget(label)
                
                # 添加提示信息
                info_label = QLabel("是否要保存结果？")
                info_label.setAlignment(Qt.AlignCenter)
                info_label.setWordWrap(True)
                layout.addWidget(info_label)

                # 添加按钮布局
                button_layout = QHBoxLayout()
                
                # 确认保存按钮
                save_button = QPushButton("保存")
                save_button.clicked.connect(lambda: [self.save_detection_data(output_dir, self.image_path), self.save_results(output_dir, stitched_img, sub_image_info)])
                
                # 取消按钮
                cancel_button = QPushButton("取消")
                cancel_button.clicked.connect(preview_window.close)

                
                button_layout.addWidget(save_button)
                button_layout.addWidget(cancel_button)
                
                layout.addLayout(button_layout)
                preview_window.setLayout(layout)
                preview_window.show()
                
                # 保存窗口引用以防止被垃圾回收
                self.preview_window = preview_window
                
            except Exception as e:
                print(f"预览窗口创建失败: {e}")

        # 结果显示
        if self.image_path:
            cv2.imshow('Analysis Result', image)
            cv2.waitKey(0)
            cv2.destroyAllWindows()
        else:
            self.main_window.result_img = image
            self.main_window.basic.display_image(image)
# ... existing code ...


if __name__ == '__main__':
    analyzer = Algorithm()
    analyzer.image_path = r'samples\tests\plate_4x6.png'
    analyzer.count()